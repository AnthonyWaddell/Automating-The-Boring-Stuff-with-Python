#! python3
''' Corects produce pricing in an Excel spreadsheet'''
import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet__by_name('Sheet')

# The produce and their updated prices
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# Loop over rows in Excel file looking for produce to update
for rowNum in range(2, sheet.max_row):
	produceName = sheet.cell(row = rowNum, column = 1).value
	# if thius is the product we are updating
	if produceName in PRICE_UPDATES:
		sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')