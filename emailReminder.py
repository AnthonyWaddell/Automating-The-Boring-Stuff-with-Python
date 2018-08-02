#! python3
''' Simple Python script to open and read through an excel file with user data
	and send reminder emails if user's have not paid their dues'''

import sys
import smtplib
import openpyxl

# Open the .xlsx file and find most recent month
wb = openpyxl.load_workbook('dueRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row = 1, column = lastCol).value

# Now check each member's payment status
# Set up dictionary for unpaid members
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
	payment = sheet.cell(row = r, column = lastCol).value
	# If they have not paid
	if payment != 'paid':
		# Add them to the dictionary of unpaid dues
		name = sheet.cell(row = r, column = 1).value
		email = sheet.cell(row = r, column = 2).value
		unpaidMembers[name] = email

print("Please enter your Gmail email address (with full @gmail.com extension:")
my_email = input()
print("Please enter your password: ")
my_pw = input()

# Log into the email email account
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(my_email my_pw)

# Send them reminder emails
for name, email in unpaidMembers.items():
	body = "Subject: %s dues unpaid. \nDear %s, \nOur Records show that you have not\
	paid dues for %s. Please make this payment as soon as possible. Thank you!" 
	%(latestMonth, name, latestMonth)
	print('Sending email to %s...' % emails)
	sendmailStatus = smtpObj.sendmail(my_email, email, body)

	if sendmailStatus != {}
	print('Error sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()